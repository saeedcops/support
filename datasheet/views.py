from django.shortcuts import render,redirect
from django.views.generic.list import ListView
from django.http import JsonResponse,HttpResponse,HttpResponseBadRequest
from core.models import *
from django.views import View
from django.db.models import F
from django.views.generic.edit import UpdateView
from django.contrib import messages
import json
from django.core.paginator import Paginator
from django import forms
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import load_workbook

class UploadFileForm(forms.Form):
    file = forms.FileField()


class UploadView(View):

    def get(self,request):

        if request.user.is_staff:
             
            return render(request, 'datasheet/upload_datasheet.html')


    def post(self,request):

        if request.user.is_staff:

            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                wb = load_workbook(filename=request.FILES['file'].file, read_only=True)
                
                pc= wb['PC']
                server= wb['Server']
                printer= wb['Printer']
                network= wb['Network']
                storage= wb['Storage']
                stock= wb['Stock']
                backup= wb['Backup']
            #PC
                for i, row in enumerate(pc.iter_rows()):
                    if i==0:
                        continue
                
                    branch=None
                    dep=None
                    try:
                        dep=Department.objects.get(name=row[1].value)
                    except ObjectDoesNotExist:
                        dep=None
                    try:
                        branch=Branch.objects.get(name=row[0].value)
                    except ObjectDoesNotExist:
                        continue
                        branch=None

                    PC.objects.update_or_create(host_name=row[2].value,defaults={
                                                'department':dep,'model':row[3].value,
                                                'cpu':row[4].value,'ram':row[5].value,'branch':branch,
                                                'hard_disk':row[6].value,'os':row[7].value,
                                                'ip':row[8].value,'mac':row[9].value})


            #Server
                for i, row in enumerate(server.iter_rows()):
                    if i==0:
                        print("\n\nServer")
                        continue

                    branch=None

                    try:
                        branch=Branch.objects.get(name=row[0].value)
                    except ObjectDoesNotExist:
                        continue
                        branch=None

                    Server.objects.update_or_create(host_name=row[1].value,defaults={
                                            'model':row[2].value,'cpu':row[3].value,'ram':row[4].value,
                                            'hard_disk':row[5].value,'os':row[6].value,'branch':branch,
                                            'ip':row[7].value,'mac':row[8].value,'role':row[9].value})

            #Printer
                for i, row in enumerate(printer.iter_rows()):
                    if i==0:
                        print("\n\nPrinter")
                        continue
                    print(str(row[2].value))
                    branch=None

                    try:
                        branch=Branch.objects.get(name=row[0].value)
                    except ObjectDoesNotExist:
                        continue
                        branch=None

                    Printer.objects.update_or_create(ip=row[2].value,defaults={
                                            'model':row[1].value,'branch':branch,
                                            'scan_share':row[3].value,'office':row[4].value})

            #network
                for i, row in enumerate(network.iter_rows()):
                    if i==0:
                        print("\n\nNetwork")
                        continue
                    print(str(row[2].value))
                    branch=None

                    try:
                        branch=Branch.objects.get(name=row[0].value)
                    except ObjectDoesNotExist:
                        continue
                        branch=None

                    Network.objects.update_or_create(ip=row[6].value,defaults={
                                        'model':row[2].value,'ports':row[3].value,'branch':branch,
                                        'username':row[4].value,'password':row[5].value,
                                        'device':row[1].value,'location':row[7].value})


            #stoke
                for i, row in enumerate(stock.iter_rows()):
                    if i==0:
                        print("\n\nStock")
                        continue
                    print(str(row[2].value))
                    branch=None

                    try:
                        branch=Branch.objects.get(name=row[0].value)
                    except ObjectDoesNotExist:
                        continue
                        branch=None

                    Stock.objects.update_or_create(branch=branch,qty=row[1].value,
                                        item=row[2].value,status=row[3].value)

            #storage
                for i, row in enumerate(storage.iter_rows()):
                    if i==0:
                        continue
                        print("\n\nStorage")
                    print(str(row[2].value))
                    branch=None

                    try:
                        branch=Branch.objects.get(name=row[0].value)
                    except ObjectDoesNotExist:
                        continue
                        branch=None

                    Storage.objects.update_or_create(ip=row[3].value,defaults={
                                        'capacity':row[2].value,'model':row[1].value,
                                        'serial_num':row[4].value,'branch':branch,})


            #backup
                for i, row in enumerate(backup.iter_rows()):
                    if i==0:
                        print("\n\nbackup")
                        continue
                    print(str(row[2].value))
                    branch=None
                    server=None

                    try:
                        server=Server.objects.get(host_name=row[3].value)
                    except ObjectDoesNotExist:
                        print("Backup not added !")
                        continue

                    try:
                        branch=Branch.objects.get(name=row[0].value)
                    except ObjectDoesNotExist:
                        continue
                        branch=None

                    Backup.objects.update_or_create(branch=branch,data=row[1].value,defaults={
                                        'descreption':row[2].value,'server':server,
                                        'internal_path':row[4].value,'internal_schadule':row[5].value,
                                        'internal_agent':row[6].value,'external_path':row[7].value,
                                        'external_schadule':row[8].value,'external_agent':row[9].value})

                messages.success(self.request, "File uploaded successfully!")
                return redirect('pc')
            else:
                messages.error(self.request, "Please select valid file!")
                return redirect('pc')




class PCUpdateView(UpdateView):
    model = PC
    context_object_name = 'pc'
    fields =  '__all__'
    template_name = 'datasheet/edit_pc.html'

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        messages.success(self.request, "PC updated successfully!")
        return redirect('pc')



class ServerUpdateView(UpdateView):
    model = Server
    context_object_name = 'server'
    fields =  '__all__'
    template_name = 'datasheet/edit_server.html'

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        messages.success(self.request, "Server updated successfully!")
        return redirect('pc')


class PrinterUpdateView(UpdateView):
    model = Printer
    context_object_name = 'printer'
    fields =  '__all__'
    template_name = 'datasheet/edit_printer.html'

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        messages.success(self.request, "Printer updated successfully!")
        return redirect('pc')


class NetworkUpdateView(UpdateView):
    model = Network
    context_object_name = 'network'
    fields =  '__all__'
    template_name = 'datasheet/edit_network.html'

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        messages.success(self.request, "Device updated successfully!")
        return redirect('pc')


class SwitchUpdateView(UpdateView):
    model = Switch
    context_object_name = 'switch'
    fields =  '__all__'
    template_name = 'datasheet/edit_switch.html'

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        messages.success(self.request, "Switch updated successfully!")
        return redirect('pc')



class FirewallUpdateView(UpdateView):
    model = Firewall
    context_object_name = 'firewall'
    fields =  '__all__'
    template_name = 'datasheet/edit_firewall.html'

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        messages.success(self.request, "Firewall updated successfully!")
        return redirect('pc')


class DVRUpdateView(UpdateView):
    model = DVR
    context_object_name = 'dvr'
    fields =  '__all__'
    template_name = 'datasheet/edit_dvr.html'

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        messages.success(self.request, "DVR updated successfully!")
        return redirect('pc')


class FingerPrintUpdateView(UpdateView):
    model = FingerPrint
    context_object_name = 'fingerprint'
    fields =  '__all__'
    template_name = 'datasheet/edit_fingerprint.html'

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        messages.success(self.request, "FingerPrint updated successfully!")
        return redirect('pc')


class PCListView(View):

    def get(self,request):

        if request.user.is_staff:
             
            ctx= PC.objects.filter(branch=request.user.branch)
            branch= Branch.objects.all()
            paginator = Paginator(ctx, 6)
            page_number = request.GET.get('page')
            page_obj = Paginator.get_page(paginator, page_number)
            return render(request, 'datasheet/index.html', {'branch': branch, 'page_obj': page_obj,'pc':ctx})


class DeviceListView(View):


    def post(self,request):

        if request.user.is_staff:

            search_str = json.loads(request.body).get('searchText')
            device = json.loads(request.body).get('device')

            if device == "PC":
                ctx= PC.objects.filter(host_name__istartswith=search_str) | \
                     PC.objects.filter(ip__istartswith=search_str) | \
                     PC.objects.filter(user__username__istartswith=search_str) | \
                     PC.objects.filter(branch__name__istartswith=search_str)

                data=ctx.values(name=F('branch__name'),username=F('user__username')).values()

            elif device == "Servers":
                ctx= Server.objects.filter(host_name__istartswith=search_str) | \
                     Server.objects.filter(ip__istartswith=search_str) | \
                     Server.objects.filter(role__istartswith=search_str) | \
                     Server.objects.filter(branch__name__istartswith=search_str)

                data=ctx.values(name=F('branch__name')).values()

            elif device == "Printers":
                ctx= Printer.objects.filter(model__istartswith=search_str) | \
                     Printer.objects.filter(ip__istartswith=search_str) | \
                     Printer.objects.filter(role__istartswith=search_str) | \
                     Printer.objects.filter(branch__name__istartswith=search_str)

                data=ctx.values(name=F('branch__name')).values()

            elif device == "Backup":
                ctx= Backup.objects.filter(data__istartswith=search_str) | \
                     Backup.objects.filter(server__host_name__istartswith=search_str) | \
                     Backup.objects.filter(internal_path__icontains=search_str) | \
                     Backup.objects.filter(external_path__icontains=search_str) 

                data=ctx.values(name=F('branch__name'),sname=F('server__host_name')).values()

            else:
                ctx= Network.objects.filter(device__istartswith=search_str) | \
                     Network.objects.filter(model__istartswith=search_str) | \
                     Network.objects.filter(location__istartswith=search_str) | \
                     Network.objects.filter(ip__istartswith=search_str)

                data=ctx.values(name=F('branch__name')).values()



            print("Data:",str(ctx.values()))
            # data=ctx.values(name=F('branch__name')).values()
            return JsonResponse(list(data),safe=False)
        else:
            return JsonResponse()




    def get(self,request):

        if request.user.is_staff:

            device = self.request.GET.get('device', 'give-default-value')
            branch = self.request.GET.get('branch', 'give-default-value')
            print("device",device)
            print("branch",branch)
            ctx= Server.objects.filter(branch__name=branch).values(name=F('branch__name'))

            if device == "PC":
                ctx= PC.objects.filter(branch__name=branch).values(name=F('branch__name'),username=F('user__username'))

            elif device == "Servers":

                ctx= Server.objects.filter(branch__name=branch).values(name=F('branch__name'))
            elif device == "Printers":

                ctx= Printer.objects.filter(branch__name=branch).values(name=F('branch__name'))

            elif device == "Switches":
                ctx= Network.objects.filter(branch__name=branch,device="Switch").values(name=F('branch__name'))

            elif device == "Firewall":
                ctx= Network.objects.filter(branch__name=branch,device="Firewall").values(name=F('branch__name'))

            elif device == "DVR":
                ctx= Network.objects.filter(branch__name=branch,device="DVR").values(name=F('branch__name'))
            
            elif device == "Router":
                ctx= Network.objects.filter(branch__name=branch,device="Router").values(name=F('branch__name'))

            elif device == "FP":
                ctx= Network.objects.filter(branch__name=branch,device="FP").values(name=F('branch__name'))

            elif device == "AP":
                ctx= Network.objects.filter(branch__name=branch,device="AP").values(name=F('branch__name'))

            elif device == "Backup":
                ctx= Backup.objects.filter(branch__name=branch).values(name=F('branch__name'),sname=F('server__host_name'))

            print("Data:",str(ctx.values()))
            # print("Branch:",str(ctx[0].branch))
            # print("branch",str(ctx[0].branch_id))
            data=ctx.values()
            return JsonResponse(list(data),safe=False)
        else:
            return JsonResponse()


class PrinterListView(ListView):

    model = Printer
    template_name = 'datasheet/index.html'
    # permission_classes = (permissions.IsAuthenticated,permissions.IsAdminUser,)
    context_object_name = 'pc'
    paginate_by = 5  # if pagination is desired